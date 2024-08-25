import ifcopenshell
import pandas as pd
import tkinter as tk
from tkinter import messagebox
from tkinter.filedialog import askopenfilename, asksaveasfilename
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def main():
    def on_submit():
        # Check if IFC file is selected
        if not ifc_file_path.get():
            messagebox.showwarning("Warning", "You must select an IFC file!")
            return

        # Check if Excel file save location is selected
        if not excel_file_path.get():
            messagebox.showwarning("Warning", "You must select a save location for the Excel file!")
            return

        # Get selected categories from the listbox
        selected_categories = [category_listbox.get(i) for i in category_listbox.curselection()]
        if not selected_categories:
            messagebox.showwarning("Warning", "You must select at least one category!")
            return
        
        # Run extraction process with selected categories
        run_combined_extraction(selected_categories, ifc_file_path.get(), excel_file_path.get())
        root.destroy()

    # Open file dialog to browse for IFC file
    def browse_ifc_file():
        file_path = askopenfilename(title="Select the IFC file", filetypes=[("IFC files", "*.ifc")])
        if file_path:
            ifc_file_path.set(file_path)
            label_ifc_file.config(text=f"IFC File: {file_path}")
            load_categories(file_path)

    # Open file dialog to choose save location for Excel file
    def browse_excel_file():
        file_path = asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")], title="Save Excel file as")
        if file_path:
            excel_file_path.set(file_path)
            label_excel_file.config(text=f"Excel File: {file_path}")

    # Load categories from the IFC file
    def load_categories(ifc_file_path):
        all_categories = load_ifc_data(ifc_file_path)
        update_category_options(all_categories)

    # Load data from the IFC file and return a sorted list of categories
    def load_ifc_data(ifc_file_path):
        ifc_file = load_ifc_file(ifc_file_path)
        if ifc_file is None:
            return []
        
        categories = set()
        # Iterate through all IFC products to collect their types
        for entity in ifc_file.by_type("IfcProduct"):
            categories.add(entity.is_a())
        categories = list(categories)
        categories.sort()

        return categories

    # Update the Listbox with available categories
    def update_category_options(categories):
        category_listbox.delete(0, tk.END)
        for category in categories:
            category_listbox.insert(tk.END, category)

    # Select all categories in the Listbox
    def on_select_all():
        category_listbox.select_set(0, tk.END)

    # Deselect all categories in the Listbox
    def on_deselect_all():
        category_listbox.selection_clear(0, tk.END)

    # Setup the main window for the application
    root = tk.Tk()
    root.title("IFC Extractor")

    border_color = '#5D8A66'

    # Variables for file paths
    ifc_file_path = tk.StringVar()
    excel_file_path = tk.StringVar()

    # Button to browse IFC file
    browse_ifc_button = tk.Button(root, text="Browse IFC File", bg="white", fg=border_color, command=browse_ifc_file, borderwidth=2)
    browse_ifc_button.pack(pady=5, padx=10)

    # Label displaying selected IFC file path
    label_ifc_file = tk.Label(root, text="IFC File: Not selected", bg='#5D8A66', fg='white')
    label_ifc_file.pack(pady=5, padx=10)

    # Button to browse for Excel file save location
    browse_excel_button = tk.Button(root, text="Browse Excel File", bg="white", fg=border_color, command=browse_excel_file, borderwidth=2)
    browse_excel_button.pack(pady=5, padx=10)

    # Label displaying selected Excel file path
    label_excel_file = tk.Label(root, text="Excel File: Not selected", bg='#5D8A66', fg='white')
    label_excel_file.pack(pady=5, padx=10)

    # Label and Listbox for category selection
    label_category = tk.Label(root, text="Select categories to include:", bg='#5D8A66', fg='white', borderwidth=2)
    label_category.pack(pady=10, padx=10)

    category_listbox = tk.Listbox(root, selectmode=tk.MULTIPLE)
    category_listbox.pack(pady=10, padx=10, fill=tk.BOTH, expand=True)

    # Button to select all categories
    select_all_button = tk.Button(root, text="Select All Categories", bg="white", fg=border_color, command=on_select_all, borderwidth=2)
    select_all_button.pack(pady=5, padx=10)

    # Button to deselect all categories
    deselect_all_button = tk.Button(root, text="Deselect All Categories", bg="white", fg=border_color, command=on_deselect_all, borderwidth=2)
    deselect_all_button.pack(pady=5, padx=10)

    # Button to start extraction
    submit_button = tk.Button(root, text="Extract", bg="white", fg=border_color, command=on_submit, borderwidth=2)
    submit_button.pack(pady=10, padx=10)

    root.geometry("400x600")
    root.configure(bg='#5D8A66')

    root.mainloop()

# Function to load the IFC file using ifcopenshell
def load_ifc_file(file_path):
    try:
        return ifcopenshell.open(file_path)
    except Exception as e:
        print(f"Error loading IFC file: {e}")
        return None

# Main function to extract and combine the data into an Excel file
def run_combined_extraction(categories, ifc_file_path, excel_file_path):
    ifc_file = load_ifc_file(ifc_file_path)
    if ifc_file is None:
        return

    writer = pd.ExcelWriter(excel_file_path, engine='openpyxl')

    # Loop through each selected category and extract its data
    for category in categories:
        category_data = extract_category_data(ifc_file, category)
        if category_data:
            df = pd.DataFrame(category_data)
            df.to_excel(writer, sheet_name=f'{category}_Properties', index=False)

    writer.close()
    customize_excel(excel_file_path)
    messagebox.showinfo("Information", f"Data saved in {excel_file_path}")

# Function to extract data from the IFC file for the selected category
def extract_category_data(ifc_file, category_name):
    elements_data = []
    elements = ifc_file.by_type(category_name)  # Filter by the selected category

    for element in elements:
        element_data = {"Category": category_name}
        if hasattr(element, 'Name'):
            element_data['Name'] = element.Name
        
        psets_data = {}
        # Extract Property Sets associated with the element
        for pset_relation in element.IsDefinedBy:
            if pset_relation.is_a("IfcRelDefinesByProperties"):
                property_set = pset_relation.RelatingPropertyDefinition
                if property_set.is_a("IfcPropertySet"):
                    pset_name = property_set.Name
                    pset_data = {}
                    for property in property_set.HasProperties:
                        if property.is_a("IfcPropertySingleValue"):
                            name = property.Name
                            value = property.NominalValue
                            if value:
                                pset_data[name] = value.wrappedValue
                        elif property.is_a("IfcPropertyEnumeratedValue"):
                            name = property.Name
                            value = property.EnumerationValues
                            if value:
                                pset_data[name] = [val.wrappedValue for val in value]
                    
                    psets_data[pset_name] = pset_data
        
        # Combine property set data into the element data
        for pset_name, pset_data in psets_data.items():
            for param_name, param_value in pset_data.items():
                element_data[f"{pset_name} - {param_name}"] = param_value

        elements_data.append(element_data)

    return elements_data

# Function to customize the appearance of the Excel file after data extraction
def customize_excel(excel_file_path):
    wb = load_workbook(excel_file_path)
    green_fill = PatternFill(start_color="5D8A66", end_color="5D8A66", fill_type="solid")

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        
        # Set the green background for column headers
        for cell in ws[1]:
            cell.fill = green_fill

        # Adjust the column width to fit content
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value:
                        value_length = len(str(cell.value))
                        if value_length > max_length:
                            max_length = value_length
                except:
                    pass
            adjusted_width = max(max_length, 10)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_file_path)

# Start the application when script is run directly
if __name__ == "__main__":
    main()